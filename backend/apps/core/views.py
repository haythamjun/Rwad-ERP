from datetime import date, timedelta

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from django.http import HttpResponse
from django.utils.dateparse import parse_date

from rest_framework import generics, serializers, status
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from django.db.models import Count, Q
from .models import AuditLog, Branch, Bus, SiteSettings, AcademicTerm, Holiday
from .permissions import CanViewReports
from .utils import log_action
from apps.accounts.permissions import IsManagerOrAbove, IsAdmin


# ── Audit Log ──────────────────────────────────────────────────────────────────

class AuditLogSerializer(serializers.ModelSerializer):
    action_display = serializers.CharField(source='get_action_display', read_only=True)
    user_name = serializers.SerializerMethodField()

    class Meta:
        model  = AuditLog
        fields = ['id', 'user', 'user_name', 'action', 'action_display',
                  'object_repr', 'changes', 'ip_address', 'timestamp']

    def get_user_name(self, obj):
        if obj.user:
            return obj.user.get_full_name() or obj.user.username
        return None


class AuditLogListView(generics.ListAPIView):
    queryset           = AuditLog.objects.select_related('user').order_by('-timestamp')
    serializer_class   = AuditLogSerializer
    permission_classes = [IsManagerOrAbove]
    filterset_fields   = ['action']
    search_fields      = ['object_repr', 'user__username']


# ── Branch ─────────────────────────────────────────────────────────────────────

class BranchSerializer(serializers.ModelSerializer):
    student_count = serializers.IntegerField(read_only=True)

    class Meta:
        model  = Branch
        fields = ['id', 'name', 'city', 'location', 'phone', 'is_active', 'created_at', 'student_count']
        read_only_fields = ['id', 'created_at']


class BranchListCreateView(generics.ListCreateAPIView):
    serializer_class   = BranchSerializer
    pagination_class   = None          # branches are a small list; no pagination needed

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get_queryset(self):
        return Branch.objects.annotate(student_count=Count('students')).order_by('name')


class BranchDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = BranchSerializer
    permission_classes = [IsAdmin]

    def get_queryset(self):
        return Branch.objects.annotate(student_count=Count('students'))


# ── إعدادات المركز ────────────────────────────────────────────────────────────

class SiteSettingsSerializer(serializers.ModelSerializer):
    class Meta:
        model  = SiteSettings
        fields = [
            'center_name_ar', 'center_name_en', 'phone', 'email',
            'website', 'address', 'logo', 'weekly_off_days', 'updated_at',
        ]
        read_only_fields = ['updated_at']


# ── الفصول الدراسية والعطل الرسمية ────────────────────────────────────────────
# (تُستخدَم لاحتساب "أيام الدراسة المتوقعة" بتقرير الحضور بشكل صحيح)

class AcademicTermSerializer(serializers.ModelSerializer):
    class Meta:
        model  = AcademicTerm
        fields = ['id', 'name', 'start_date', 'end_date', 'created_at']
        read_only_fields = ['id', 'created_at']


class AcademicTermListCreateView(generics.ListCreateAPIView):
    serializer_class = AcademicTermSerializer
    queryset         = AcademicTerm.objects.all()
    pagination_class = None

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdmin()]
        return [IsAuthenticated()]


class AcademicTermDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = AcademicTermSerializer
    queryset           = AcademicTerm.objects.all()
    permission_classes = [IsAdmin]


class HolidaySerializer(serializers.ModelSerializer):
    class Meta:
        model  = Holiday
        fields = ['id', 'name', 'start_date', 'end_date', 'created_at']
        read_only_fields = ['id', 'created_at']


class HolidayListCreateView(generics.ListCreateAPIView):
    serializer_class = HolidaySerializer
    queryset         = Holiday.objects.all()
    pagination_class = None

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdmin()]
        return [IsAuthenticated()]


class HolidayDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = HolidaySerializer
    queryset           = Holiday.objects.all()
    permission_classes = [IsAdmin]


class SiteSettingsView(APIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_permissions(self):
        if self.request.method == 'PATCH':
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get(self, request):
        return Response(SiteSettingsSerializer(SiteSettings.get_solo()).data)

    def patch(self, request):
        obj = SiteSettings.get_solo()
        serializer = SiteSettingsSerializer(obj, data=request.data, partial=True)
        serializer.is_valid(raise_exception=True)
        serializer.save()
        return Response(serializer.data)


# ── Bus ────────────────────────────────────────────────────────────────────────

class BusSerializer(serializers.ModelSerializer):
    branch_name = serializers.CharField(source='branch.name', read_only=True)

    class Meta:
        model  = Bus
        fields = [
            'id', 'chassis_number', 'plate_number', 'brand', 'manufacture_year',
            'serial_number', 'branch', 'branch_name',
            'registration_expiry', 'inspection_expiry',
            'created_at', 'updated_at',
        ]
        read_only_fields = ['id', 'created_at', 'updated_at']

    def validate_manufacture_year(self, value):
        current_year = date.today().year
        if not (1980 <= value <= current_year + 1):
            raise serializers.ValidationError(f'سنة الصنع يجب أن تكون بين 1980 و {current_year + 1}.')
        return value


class BusListCreateView(generics.ListCreateAPIView):
    serializer_class = BusSerializer
    pagination_class = None  # عدد الباصات محدود عادة — لا حاجة للترقيم
    filterset_fields = ['branch']

    def get_permissions(self):
        if self.request.method == 'POST':
            return [IsAdmin()]
        return [IsAuthenticated()]

    def get_queryset(self):
        return Bus.objects.select_related('branch').order_by('-created_at')


class BusDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = BusSerializer
    permission_classes = [IsAdmin]
    queryset           = Bus.objects.select_related('branch')


# ── Dashboard stats ────────────────────────────────────────────────────────────

class DashboardStatsView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        from apps.students.models import Student
        from apps.accounts.models import User

        user = request.user
        qs = Student.objects.all()
        branches_qs = Branch.objects.filter(is_active=True)
        buses_qs = Bus.objects.all()

        if not user.is_admin:
            if user.assigned_branch_id:
                qs = qs.filter(branch_id=user.assigned_branch_id)
                branches_qs = branches_qs.filter(id=user.assigned_branch_id)
                buses_qs = buses_qs.filter(branch_id=user.assigned_branch_id)
            elif user.assigned_city:
                qs = qs.filter(branch__city=user.assigned_city)
                branches_qs = branches_qs.filter(city=user.assigned_city)
                buses_qs = buses_qs.filter(branch__city=user.assigned_city)

        total = qs.count()
        by_status = dict(
            qs.values_list('status').annotate(n=Count('id')).values_list('status', 'n')
        )

        branches = (
            branches_qs
            .annotate(
                student_count=Count('students', distinct=True),
                male_count=Count('students', filter=Q(students__gender='male'), distinct=True),
                female_count=Count('students', filter=Q(students__gender='female'), distinct=True),
            )
            .values('id', 'name', 'location', 'student_count', 'male_count', 'female_count')
            .order_by('name')
        )

        today = date.today()
        bus_count = buses_qs.count()
        buses_needing_renewal = buses_qs.filter(
            Q(registration_expiry__lt=today) | Q(inspection_expiry__lt=today)
        ).count()

        # عدد المستخدمين — نطاق عام (نظام) وليس محصورًا بالفرع
        users_qs = User.objects.filter(is_active=True)
        user_count = users_qs.count()
        users_by_role = dict(
            users_qs.values_list('role').annotate(n=Count('id')).values_list('role', 'n')
        )

        return Response({
            'total':     total,
            'by_status': by_status,
            'branches':  list(branches),
            'bus_count': bus_count,
            'buses_needing_renewal': buses_needing_renewal,
            'user_count': user_count,
            'users_by_role': users_by_role,
        })


# ── التقارير — تقرير الحضور والانصراف ────────────────────────────────────────

_WEEKDAY_NAMES = ['monday', 'tuesday', 'wednesday', 'thursday', 'friday', 'saturday', 'sunday']


def _expected_school_days(date_from, date_to):
    """يُرجع set() من تواريخ أيام الدراسة "المتوقّعة" ضمن [date_from, date_to] —
    مستبعدًا أيام الإجازة الأسبوعية (من الإعدادات) والعطل الرسمية، ومقصورًا على
    نطاق الفصول الدراسية المعرَّفة. لو لم يُعرَّف أي فصل دراسي بالنظام إطلاقًا
    بعد، لا يُطبَّق شرط "داخل فصل" (تفاديًا لنسبة 0% مضلِّلة قبل إعداد الفصول)."""
    off_days = set(SiteSettings.get_solo().weekly_off_days or [])

    holiday_dates = set()
    for h in Holiday.objects.filter(start_date__lte=date_to, end_date__gte=date_from):
        d = max(h.start_date, date_from)
        end = min(h.end_date, date_to)
        while d <= end:
            holiday_dates.add(d)
            d += timedelta(days=1)

    terms = list(AcademicTerm.objects.filter(start_date__lte=date_to, end_date__gte=date_from))
    any_term_defined = AcademicTerm.objects.exists()

    def in_some_term(d):
        return True if not any_term_defined else any(t.start_date <= d <= t.end_date for t in terms)

    expected = set()
    d = date_from
    while d <= date_to:
        if _WEEKDAY_NAMES[d.weekday()] not in off_days and d not in holiday_dates and in_some_term(d):
            expected.add(d)
        d += timedelta(days=1)
    return expected


def _build_attendance_report(request):
    """يبني بيانات تقرير الحضور: (date_from, date_to, summary, rows) — يُستخدم من عرض JSON وتصدير Excel معًا.

    النسبة تُحتسب من "أيام الدراسة المتوقعة" الفعلية (وليس عدد السجلات الموجودة
    فقط) — أي يوم متوقّع بلا سجل حضور إطلاقًا يُحتسب غيابًا. التفصيل حسب الطالب
    يشمل كل طالب نشط ضمن النطاق حتى لو لم يُسجَّل له أي حضور بالفترة."""
    from apps.students.models import Student, StudentAttendance

    today = date.today()
    date_from = parse_date(request.query_params.get('date_from', '')) or today.replace(day=1)
    date_to   = parse_date(request.query_params.get('date_to', '')) or today

    expected_dates = _expected_school_days(date_from, date_to)
    expected_days  = len(expected_dates)

    # أي يوم بالمدى ليس من أيام الدراسة المتوقعة (عطلة أسبوعية/رسمية أو خارج كل
    # الفصول) — تُستبعد سجلات الحضور الواقعة عليه من الحساب (شذوذ نادر).
    excluded_dates = []
    d = date_from
    while d <= date_to:
        if d not in expected_dates:
            excluded_dates.append(d)
        d += timedelta(days=1)

    qs = StudentAttendance.objects.filter(attendance_date__gte=date_from, attendance_date__lte=date_to)
    if excluded_dates:
        qs = qs.exclude(attendance_date__in=excluded_dates)
    students_qs = Student.objects.filter(status='active')

    user = request.user
    if not user.is_admin:
        if user.assigned_branch_id:
            qs = qs.filter(branch_id=user.assigned_branch_id)
            students_qs = students_qs.filter(branch_id=user.assigned_branch_id)
        elif user.assigned_city:
            qs = qs.filter(branch__city=user.assigned_city)
            students_qs = students_qs.filter(branch__city=user.assigned_city)

    branch_param = request.query_params.get('branch')
    if branch_param:
        qs = qs.filter(branch_id=branch_param)
        students_qs = students_qs.filter(branch_id=branch_param)

    student_param = request.query_params.get('student')
    if student_param:
        qs = qs.filter(student_id=student_param)
        students_qs = students_qs.filter(id=student_param)

    student_count = students_qs.count()

    def _rate(present, late, denom):
        return round((present + late) / denom * 100, 1) if denom else 0

    # ── الملخص العام ──────────────────────────────────────────────────────────
    marked = qs.aggregate(
        present=Count('id', filter=Q(status='present')),
        late=Count('id', filter=Q(status='late')),
        excused_absence=Count('id', filter=Q(status='excused_absence')),
        early_leave=Count('id', filter=Q(status='early_leave')),
    )
    total_slots = expected_days * student_count
    absent = max(total_slots - marked['present'] - marked['late'] - marked['excused_absence'] - marked['early_leave'], 0)
    summary = {
        'total':            total_slots,
        'present':          marked['present'],
        'absent':           absent,
        'late':             marked['late'],
        'excused_absence':  marked['excused_absence'],
        'early_leave':      marked['early_leave'],
        'attendance_rate':  _rate(marked['present'], marked['late'], total_slots),
        'student_count':    student_count,
        'expected_days':    expected_days,
    }

    # ── التفصيل حسب الطالب ───────────────────────────────────────────────────
    ZERO_STATS = {'present': 0, 'late': 0, 'excused_absence': 0, 'early_leave': 0}
    stats_by_student = {
        r['student_id']: r
        for r in qs.values('student_id').annotate(
            present=Count('id', filter=Q(status='present')),
            late=Count('id', filter=Q(status='late')),
            excused_absence=Count('id', filter=Q(status='excused_absence')),
            early_leave=Count('id', filter=Q(status='early_leave')),
        )
    }

    rows = []
    for s in students_qs.select_related('branch'):
        stat = stats_by_student.get(s.id, ZERO_STATS)
        present, late = stat['present'], stat['late']
        excused, early = stat['excused_absence'], stat['early_leave']
        student_absent = max(expected_days - present - late - excused - early, 0)
        rows.append({
            'student_id':       s.id,
            'student_name':     s.full_name,
            'file_number':      s.file_number,
            'branch_name':      s.branch.name if s.branch_id else None,
            'total':            expected_days,
            'present':          present,
            'absent':           student_absent,
            'late':             late,
            'excused_absence':  excused,
            'early_leave':      early,
            'attendance_rate':  _rate(present, late, expected_days),
        })
    rows.sort(key=lambda r: r['student_name'])

    return date_from, date_to, summary, rows


class AttendanceReportView(APIView):
    """تقرير الحضور والانصراف — إحصائيات مجمّعة عبر مدى تاريخي (وليس تسجيلًا يوميًا)."""
    permission_classes = [CanViewReports]

    def get(self, request):
        date_from, date_to, summary, rows = _build_attendance_report(request)
        return Response({
            'date_from': date_from,
            'date_to':   date_to,
            'summary':   summary,
            'by_student': rows,
        })


class AttendanceReportExportView(APIView):
    permission_classes = [CanViewReports]

    def get(self, request):
        date_from, date_to, summary, rows = _build_attendance_report(request)

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'تقرير الحضور'
        ws.sheet_view.rightToLeft = True

        title_fill = PatternFill(start_color='0F2A47', end_color='0F2A47', fill_type='solid')
        title_font = Font(color='FFFFFF', bold=True, size=13)
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=10)
        cell = ws.cell(row=1, column=1, value=f'تقرير الحضور والانصراف — من {date_from} إلى {date_to}')
        cell.fill = title_fill
        cell.font = title_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 24

        headers = ['رقم الملف', 'اسم الطالب', 'الفرع', 'حاضر', 'غائب', 'متأخر', 'غياب بعذر', 'انصراف مبكر', 'أيام الدراسة المتوقعة', 'نسبة الحضور %']
        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        ws.row_dimensions[2].height = 22
        for col_idx, header in enumerate(headers, 1):
            c = ws.cell(row=2, column=col_idx, value=header)
            c.fill = header_fill
            c.font = header_font
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[c.column_letter].width = 16

        alt_fill = PatternFill(start_color='EEF2F8', end_color='EEF2F8', fill_type='solid')
        for row_idx, r in enumerate(rows, 3):
            values = [
                r['file_number'], r['student_name'], r['branch_name'] or '',
                r['present'], r['absent'], r['late'], r['excused_absence'], r['early_leave'],
                r['total'], r['attendance_rate'],
            ]
            for col_idx, value in enumerate(values, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=value)
                c.alignment = Alignment(horizontal='center', vertical='center')
                if row_idx % 2 == 0:
                    c.fill = alt_fill

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="attendance_report_{date_from}_{date_to}.xlsx"'
        wb.save(response)

        log_action(request, 'export', None, f'تصدير تقرير الحضور والانصراف ({date_from} إلى {date_to})')
        return response
