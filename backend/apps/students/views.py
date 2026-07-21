from rest_framework import generics, status, filters
from rest_framework import serializers as drf_serializers
from rest_framework.views import APIView
from rest_framework.response import Response
from rest_framework.parsers import MultiPartParser, FormParser, JSONParser
from rest_framework.permissions import IsAuthenticated
from django_filters.rest_framework import DjangoFilterBackend
from django.http import HttpResponse
from django.shortcuts import get_object_or_404

import csv
import io

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

from .models import Student, Guardian, FamilyInfo, StudentAttachment
from .serializers import (
    StudentListSerializer,
    StudentDetailSerializer,
    StudentCreateUpdateSerializer,
    GuardianSerializer,
    FamilyInfoSerializer,
    StudentAttachmentSerializer,
)
from .filters import StudentFilter
from .permissions import CanWrite, CanDelete, CanExport, CanImport
from apps.core.utils import log_action


def _disability_display(student):
    """Comma-joined Arabic labels for a student's disability_type list."""
    type_map = dict(Student.DisabilityType.choices)
    types = student.disability_type if isinstance(student.disability_type, list) else []
    return '، '.join(type_map.get(t, t) for t in types if t)


def _apply_scope(user, qs):
    """Restrict queryset to the user's assigned branch or city. Admin sees all."""
    if user.is_admin:
        return qs
    if user.assigned_branch_id:
        return qs.filter(branch_id=user.assigned_branch_id)
    if user.assigned_city:
        return qs.filter(branch__city=user.assigned_city)
    return qs


class StudentListCreateView(generics.ListCreateAPIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]
    filter_backends = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_class = StudentFilter
    search_fields = ['first_name', 'middle_name', 'grandfather_name', 'family_name', 'national_id', 'file_number']
    ordering_fields = ['first_name', 'family_name', 'registration_date', 'created_at', 'file_number']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = Student.objects.select_related('created_by').prefetch_related('guardians', 'family_info')
        return _apply_scope(self.request.user, qs)

    def get_serializer_class(self):
        if self.request.method == 'POST':
            return StudentCreateUpdateSerializer
        return StudentListSerializer

    def get_permissions(self):
        if self.request.method == 'POST':
            return [CanWrite()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        student = serializer.save()
        log_action(self.request, 'create', student, str(student))


class StudentDetailView(generics.RetrieveUpdateDestroyAPIView):
    parser_classes = [MultiPartParser, FormParser, JSONParser]

    def get_queryset(self):
        qs = Student.objects.select_related('created_by').prefetch_related('guardians', 'family_info', 'attachments')
        return _apply_scope(self.request.user, qs)

    def get_serializer_class(self):
        if self.request.method in ['PUT', 'PATCH']:
            return StudentCreateUpdateSerializer
        return StudentDetailSerializer

    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH']:
            return [CanWrite()]
        if self.request.method == 'DELETE':
            return [CanDelete()]
        return [IsAuthenticated()]

    def perform_update(self, serializer):
        student = serializer.save()
        log_action(self.request, 'update', student, str(student))

    def perform_destroy(self, instance):
        log_action(self.request, 'delete', instance, str(instance))
        instance.delete()


# ──────────────────────────────────────────────────────────────────────────────
# Reject student
# ──────────────────────────────────────────────────────────────────────────────

class RejectStudentView(APIView):
    permission_classes = [CanDelete]

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        reason = (request.data.get('reason') or '').strip()
        if not reason:
            return Response(
                {'reason': ['سبب الرفض مطلوب']},
                status=status.HTTP_400_BAD_REQUEST,
            )
        student.status = Student.Status.REJECTED
        student.rejection_reason = reason
        student.save(update_fields=['status', 'rejection_reason', 'updated_at'])
        log_action(request, 'update', student, f'رفض الطالب: {student}')
        return Response({'status': 'rejected', 'rejection_reason': reason})


class RestoreStudentView(APIView):
    permission_classes = [CanDelete]

    def post(self, request, pk):
        student = get_object_or_404(Student, pk=pk)
        student.status = Student.Status.PENDING
        student.rejection_reason = ''
        student.save(update_fields=['status', 'rejection_reason', 'updated_at'])
        log_action(request, 'update', student, f'استعادة الطالب: {student}')
        return Response({'status': 'pending'})


# ──────────────────────────────────────────────────────────────────────────────
# Guardian views
# ──────────────────────────────────────────────────────────────────────────────

class GuardianListCreateView(generics.ListCreateAPIView):
    serializer_class = GuardianSerializer

    def get_queryset(self):
        return Guardian.objects.filter(student_id=self.kwargs['student_pk'])

    def get_permissions(self):
        if self.request.method == 'POST':
            return [CanWrite()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        student = get_object_or_404(Student, pk=self.kwargs['student_pk'])
        guardian = serializer.save(student=student)
        log_action(self.request, 'create', guardian, str(guardian))


class GuardianDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class = GuardianSerializer

    def get_queryset(self):
        return Guardian.objects.filter(student_id=self.kwargs['student_pk'])

    def get_permissions(self):
        if self.request.method in ['PUT', 'PATCH']:
            return [CanWrite()]
        if self.request.method == 'DELETE':
            return [CanDelete()]
        return [IsAuthenticated()]

    def perform_update(self, serializer):
        guardian = serializer.save()
        log_action(self.request, 'update', guardian, str(guardian))

    def perform_destroy(self, instance):
        log_action(self.request, 'delete', instance, str(instance))
        instance.delete()


# ──────────────────────────────────────────────────────────────────────────────
# FamilyInfo views
# ──────────────────────────────────────────────────────────────────────────────

class FamilyInfoView(APIView):
    def get_permissions(self):
        if self.request.method in ['POST', 'PUT', 'PATCH']:
            return [CanWrite()]
        return [IsAuthenticated()]

    def get(self, request, student_pk):
        student = get_object_or_404(Student, pk=student_pk)
        try:
            family = student.family_info
        except FamilyInfo.DoesNotExist:
            return Response(
                {'detail': 'لا توجد بيانات أسرة لهذا الطالب بعد'},
                status=status.HTTP_404_NOT_FOUND,
            )
        serializer = FamilyInfoSerializer(family)
        return Response(serializer.data)

    def post(self, request, student_pk):
        student = get_object_or_404(Student, pk=student_pk)
        if hasattr(student, 'family_info'):
            return Response(
                {'detail': 'يوجد بالفعل سجل أسرة لهذا الطالب. استخدم PUT للتعديل.'},
                status=status.HTTP_400_BAD_REQUEST,
            )
        serializer = FamilyInfoSerializer(data=request.data)
        if serializer.is_valid():
            family = serializer.save(student=student)
            log_action(request, 'create', family, str(family))
            return Response(serializer.data, status=status.HTTP_201_CREATED)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

    def put(self, request, student_pk):
        student = get_object_or_404(Student, pk=student_pk)
        try:
            family = student.family_info
        except FamilyInfo.DoesNotExist:
            return Response(
                {'detail': 'لا توجد بيانات أسرة. استخدم POST لإنشائها أولاً.'},
                status=status.HTTP_404_NOT_FOUND,
            )
        serializer = FamilyInfoSerializer(family, data=request.data, partial=True)
        if serializer.is_valid():
            family = serializer.save()
            log_action(request, 'update', family, str(family))
            return Response(serializer.data)
        return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


# ──────────────────────────────────────────────────────────────────────────────
# Attachment views
# ──────────────────────────────────────────────────────────────────────────────

class StudentAttachmentListView(generics.ListCreateAPIView):
    serializer_class = StudentAttachmentSerializer
    parser_classes = [MultiPartParser, FormParser]

    def get_queryset(self):
        return StudentAttachment.objects.filter(student_id=self.kwargs['student_pk'])

    def get_permissions(self):
        if self.request.method == 'POST':
            return [CanWrite()]
        return [IsAuthenticated()]

    def perform_create(self, serializer):
        student = get_object_or_404(Student, pk=self.kwargs['student_pk'])
        attachment = serializer.save(student=student, uploaded_by=self.request.user)
        log_action(self.request, 'create', attachment, str(attachment))


class StudentAttachmentDeleteView(generics.DestroyAPIView):
    serializer_class = StudentAttachmentSerializer
    permission_classes = [CanDelete]

    def get_queryset(self):
        return StudentAttachment.objects.filter(student_id=self.kwargs['student_pk'])

    def perform_destroy(self, instance):
        log_action(self.request, 'delete', instance, str(instance))
        instance.file.delete(save=False)
        instance.delete()


# ──────────────────────────────────────────────────────────────────────────────
# Export
# ──────────────────────────────────────────────────────────────────────────────

class StudentExportView(APIView):
    permission_classes = [CanExport]

    def get(self, request):
        students = _apply_scope(
            request.user,
            Student.objects.all()
            .prefetch_related('guardians')
            .select_related('family_info')
            .order_by('file_number'),
        )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'بيانات المستفيدين'
        ws.sheet_view.rightToLeft = True

        # ── Group headers (row 1) ──────────────────────────────────────
        GROUP_HEADERS = [
            ('بيانات المستفيد',          1,  15),
            ('بيانات ولي الأمر',         16, 24),
            ('دراسة الحالة الاجتماعية',  25, 31),
        ]
        group_fill = PatternFill(start_color='0F2A47', end_color='0F2A47', fill_type='solid')
        group_font = Font(color='FFFFFF', bold=True, size=13)
        for title, start_col, end_col in GROUP_HEADERS:
            ws.merge_cells(start_row=1, start_column=start_col,
                           end_row=1,   end_column=end_col)
            cell = ws.cell(row=1, column=start_col, value=title)
            cell.fill      = group_fill
            cell.font      = group_font
            cell.alignment = Alignment(horizontal='center', vertical='center')

        # ── Column headers (row 2) ─────────────────────────────────────
        headers = [
            # بيانات المستفيد
            'رقم الملف', 'تاريخ التسجيل', 'الاسم الأول', 'اسم الأب', 'اسم الجد', 'اسم العائلة',
            'رقم الهوية', 'تاريخ الميلاد', 'الجنس', 'الجنسية', 'الحالة',
            'نوع الإعاقة', 'درجة الإعاقة', 'التشخيص', 'جهة الإحالة',
            # بيانات ولي الأمر
            'اسم ولي الأمر', 'صلة القرابة', 'رقم هوية ولي الأمر',
            'رقم الجوال', 'جوال إضافي', 'البريد الإلكتروني',
            'العنوان', 'جهة التواصل الرئيسية', 'ملاحظات ولي الأمر',
            # دراسة الحالة
            'عدد أفراد الأسرة', 'ترتيب المستفيد', 'حالة الوالدين',
            'الدخل الشهري (ريال)', 'الدخل التقريبي', 'نوع السكن',
            'ملاحظات اجتماعية',
        ]

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        ws.row_dimensions[1].height = 22
        ws.row_dimensions[2].height = 22

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.fill      = header_fill
            cell.font      = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 17

        # ── Data rows ─────────────────────────────────────────────────
        alt_fill = PatternFill(start_color='EEF2F8', end_color='EEF2F8', fill_type='solid')

        for row_idx, student in enumerate(students, 3):
            guardian = (
                student.guardians.filter(is_primary_contact=True).first()
                or student.guardians.first()
            )
            fam = getattr(student, 'family_info', None)
            row_fill = alt_fill if row_idx % 2 == 0 else None

            data = [
                # بيانات المستفيد
                student.file_number,
                str(student.registration_date),
                student.first_name,
                student.middle_name,
                student.grandfather_name,
                student.family_name,
                student.national_id,
                str(student.date_of_birth),
                student.get_gender_display(),
                student.nationality,
                student.get_status_display(),
                _disability_display(student),
                student.get_disability_degree_display() if student.disability_degree else '',
                student.diagnosis or '',
                student.get_referral_source_display()  if student.referral_source  else '',
                # بيانات ولي الأمر
                guardian.full_name          if guardian else '',
                guardian.get_relationship_display() if guardian else '',
                guardian.national_id        if guardian else '',
                guardian.phone              if guardian else '',
                guardian.phone_alt          if guardian else '',
                guardian.email              if guardian else '',
                guardian.address            if guardian else '',
                'نعم' if (guardian and guardian.is_primary_contact) else 'لا',
                guardian.notes              if guardian else '',
                # دراسة الحالة
                fam.family_size    if fam else '',
                fam.sibling_order  if fam else '',
                fam.get_parents_status_display() if fam else '',
                fam.monthly_income if fam else '',
                fam.get_income_range_display()   if fam else '',
                fam.get_housing_type_display()   if fam else '',
                fam.social_notes   if fam else '',
            ]
            for col_idx, value in enumerate(data, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.alignment = Alignment(horizontal='right', vertical='center')
                if row_fill:
                    cell.fill = row_fill

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="students_roya.xlsx"'
        wb.save(response)

        log_action(request, 'export', None, 'تصدير قائمة الطلاب إلى Excel')
        return response


# ──────────────────────────────────────────────────────────────────────────────
# Import
# ──────────────────────────────────────────────────────────────────────────────

GENDER_MAP = {'ذكر': 'male', 'أنثى': 'female', 'male': 'male', 'female': 'female'}

STATUS_MAP = {
    'في انتظار القبول': 'pending',    'pending':     'pending',
    'نشط':              'active',     'active':      'active',
    'غير نشط':          'inactive',   'inactive':    'inactive',
    'خرّيج':            'graduated',  'graduated':   'graduated',
    'موقوف':            'suspended',  'suspended':   'suspended',
    'محوّل':            'transferred','transferred': 'transferred',
    'مرفوض':            'rejected',   'rejected':    'rejected',
}

DISABILITY_MAP = {
    'إعاقة ذهنية': 'intellectual', 'intellectual': 'intellectual',
    'طيف التوحد': 'autism', 'autism': 'autism',
    'متلازمة داون': 'down', 'down': 'down',
    'إعاقة حركية': 'physical', 'physical': 'physical',
    'إعاقة سمعية': 'hearing', 'hearing': 'hearing',
    'إعاقة بصرية': 'visual', 'visual': 'visual',
    'إعاقة لغوية / نطقية': 'speech', 'speech': 'speech',
    'صعوبات تعلم': 'learning', 'learning': 'learning',
    'اضطراب سلوكي': 'behavioral', 'behavioral': 'behavioral',
    'إعاقة مركّبة': 'multiple', 'multiple': 'multiple',
    'أخرى': 'other', 'other': 'other',
}

DEGREE_MAP = {
    'بسيطة': 'mild', 'mild': 'mild',
    'متوسطة': 'moderate', 'moderate': 'moderate',
    'شديدة': 'severe', 'severe': 'severe',
    'شديدة جداً': 'profound', 'profound': 'profound',
}

EDU_MAP = {
    'لا يتعلم': 'none', 'none': 'none',
    'رياض أطفال': 'kindergarten', 'kindergarten': 'kindergarten',
    'ابتدائي': 'elementary', 'elementary': 'elementary',
    'متوسط': 'intermediate', 'intermediate': 'intermediate',
    'ثانوي': 'secondary', 'secondary': 'secondary',
    'جامعي': 'university', 'university': 'university',
    'برنامج تربية خاصة': 'special', 'special': 'special',
}

REFERRAL_MAP = {
    'مستشفى / عيادة': 'hospital', 'hospital': 'hospital',
    'مدرسة': 'school', 'school': 'school',
    'الأسرة مباشرة': 'family', 'family': 'family',
    'جمعية / مؤسسة': 'ngo', 'ngo': 'ngo',
    'وزارة / جهة حكومية': 'ministry', 'ministry': 'ministry',
    'طبيب / معالج': 'specialist', 'specialist': 'specialist',
    'أخرى': 'other', 'other': 'other',
}


def _cell(row, idx):
    val = row[idx].value
    return str(val).strip() if val is not None else ''


class StudentImportView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [CanImport]

    def post(self, request):
        from datetime import date as date_type

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'لم يتم إرفاق ملف.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            wb = openpyxl.load_workbook(file, data_only=True)
            ws = wb.active
        except Exception:
            return Response({'detail': 'تعذّر قراءة الملف. تأكد أنه ملف Excel صحيح.'}, status=status.HTTP_400_BAD_REQUEST)

        rows = list(ws.iter_rows(min_row=2))
        if not rows:
            return Response({'detail': 'الملف فارغ.'}, status=status.HTTP_400_BAD_REQUEST)

        created, skipped, errors = 0, 0, []

        for i, row in enumerate(rows, start=2):
            if len(row) < 8:
                continue

            first_name       = _cell(row, 0)
            middle_name      = _cell(row, 1)
            grandfather_name = _cell(row, 2)
            family_name      = _cell(row, 3)
            national_id      = _cell(row, 4)
            dob_raw          = _cell(row, 5)
            gender_raw       = _cell(row, 6)
            nationality      = _cell(row, 7)

            if not any([first_name, family_name, national_id, dob_raw]):
                continue  # صف فارغ

            row_errors = []

            if not first_name:
                row_errors.append('الاسم الأول مطلوب')
            if not family_name:
                row_errors.append('اسم العائلة مطلوب')
            if not national_id:
                row_errors.append('رقم الهوية مطلوب')
            elif not national_id.isdigit() or len(national_id) != 10:
                row_errors.append('رقم الهوية يجب أن يكون 10 أرقام')
            if not nationality:
                row_errors.append('الجنسية مطلوبة')

            gender = GENDER_MAP.get(gender_raw)
            if not gender:
                row_errors.append(f'الجنس غير صحيح: "{gender_raw}" — استخدم: ذكر أو أنثى')

            try:
                if isinstance(row[5].value, date_type):
                    dob = row[5].value
                else:
                    from datetime import datetime
                    dob = datetime.strptime(dob_raw, '%Y-%m-%d').date()
                today = date_type.today()
                if dob > today:
                    row_errors.append('تاريخ الميلاد في المستقبل')
                elif dob < today.replace(year=today.year - 100):
                    row_errors.append('تاريخ الميلاد أكبر من 100 سنة')
            except (ValueError, TypeError):
                row_errors.append(f'تاريخ الميلاد غير صحيح: "{dob_raw}" — الصيغة المطلوبة: YYYY-MM-DD')
                dob = None

            full_name_display = f"{first_name} {family_name}".strip()

            if row_errors:
                errors.append({'row': i, 'name': full_name_display or '—', 'errors': row_errors})
                skipped += 1
                continue

            if Student.objects.filter(national_id=national_id).exists():
                errors.append({'row': i, 'name': full_name_display, 'errors': [f'رقم الهوية {national_id} مسجّل مسبقاً']})
                skipped += 1
                continue

            status_val        = STATUS_MAP.get(_cell(row, 8) if len(row) > 8 else '', 'pending') or 'pending'
            reg_date_raw      = _cell(row, 9) if len(row) > 9 else ''
            disability_raw    = _cell(row, 10) if len(row) > 10 else ''
            degree_raw        = _cell(row, 11) if len(row) > 11 else ''
            diagnosis         = _cell(row, 12) if len(row) > 12 else ''
            edu_raw           = _cell(row, 13) if len(row) > 13 else ''
            school_name       = _cell(row, 14) if len(row) > 14 else ''
            grade             = _cell(row, 15) if len(row) > 15 else ''
            referral_raw      = _cell(row, 16) if len(row) > 16 else ''
            notes             = _cell(row, 17) if len(row) > 17 else ''

            from datetime import date as date_type2
            try:
                if len(row) > 9 and isinstance(row[9].value, date_type2):
                    reg_date = row[9].value
                elif reg_date_raw:
                    from datetime import datetime
                    reg_date = datetime.strptime(reg_date_raw, '%Y-%m-%d').date()
                else:
                    reg_date = date_type2.today()
            except (ValueError, TypeError):
                reg_date = date_type2.today()

            Student.objects.create(
                first_name        = first_name,
                middle_name       = middle_name,
                grandfather_name  = grandfather_name,
                family_name       = family_name,
                national_id       = national_id,
                date_of_birth     = dob,
                gender            = gender,
                nationality       = nationality,
                status            = status_val,
                registration_date = reg_date,
                disability_type   = [DISABILITY_MAP[t.strip()] for t in disability_raw.split(',') if t.strip() in DISABILITY_MAP],
                disability_degree = DEGREE_MAP.get(degree_raw, ''),
                diagnosis         = diagnosis,
                educational_level = EDU_MAP.get(edu_raw, ''),
                school_name       = school_name,
                grade             = grade,
                referral_source   = REFERRAL_MAP.get(referral_raw, ''),
                notes             = notes,
                created_by        = request.user,
            )
            created += 1

        log_action(request, 'create', None, f'استيراد {created} طالب من Excel')
        return Response({
            'created': created,
            'skipped': skipped,
            'errors':  errors,
        }, status=status.HTTP_200_OK)


class StudentImportTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'قالب استيراد الطلاب'
        ws.sheet_view.rightToLeft = True

        headers = [
            'الاسم الأول *',
            'اسم الأب *',
            'اسم الجد',
            'اسم العائلة *',
            'رقم الهوية * (10 أرقام)',
            'تاريخ الميلاد * (YYYY-MM-DD)',
            'الجنس * (ذكر/أنثى)',
            'الجنسية *',
            'الحالة (نشط / في انتظار القبول / غير نشط / خرّيج / موقوف / محوّل)',
            'تاريخ التسجيل (YYYY-MM-DD)',
            'نوع الإعاقة (إعاقة ذهنية / طيف التوحد / متلازمة داون / إعاقة حركية / إعاقة سمعية / إعاقة بصرية / إعاقة لغوية / نطقية / صعوبات تعلم / اضطراب سلوكي / إعاقة مركّبة / أخرى)',
            'درجة الإعاقة (بسيطة / متوسطة / شديدة / شديدة جداً)',
            'التشخيص التفصيلي',
            'المستوى التعليمي (لا يتعلم / رياض أطفال / ابتدائي / متوسط / ثانوي / جامعي / برنامج تربية خاصة)',
            'اسم المدرسة',
            'الصف / المرحلة',
            'جهة الإحالة (مستشفى / عيادة / مدرسة / الأسرة مباشرة / جمعية / مؤسسة / وزارة / جهة حكومية / طبيب / معالج / أخرى)',
            'ملاحظات',
        ]

        header_fill = PatternFill(start_color='1E3A5F', end_color='1E3A5F', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=11)
        ws.row_dimensions[1].height = 30

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            ws.column_dimensions[cell.column_letter].width = 22

        # صف مثال
        example = [
            'محمد', 'أحمد', 'سعد', 'العتيبي', '1234567890', '2010-05-15', 'ذكر', 'سعودي',
            'نشط', '2024-01-10', 'طيف التوحد', 'متوسطة', 'تشخيص طيف التوحد درجة 2',
            'برنامج تربية خاصة', 'مدرسة الأمل', 'الثالث الابتدائي', 'مستشفى / عيادة', '',
        ]
        ex_fill = PatternFill(start_color='EEF2F8', end_color='EEF2F8', fill_type='solid')
        for col, val in enumerate(example, 1):
            cell = ws.cell(row=2, column=col, value=val)
            cell.fill = ex_fill
            cell.alignment = Alignment(horizontal='right', vertical='center')

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="import_template_roya.xlsx"'
        wb.save(response)
        return response


# ──────────────────────────────────────────────────────────────────────────────
# CSV Export / Import
# ──────────────────────────────────────────────────────────────────────────────

class StudentExportCsvView(APIView):
    permission_classes = [CanExport]

    def get(self, request):
        students = _apply_scope(
            request.user,
            Student.objects.all()
            .prefetch_related('guardians')
            .select_related('family_info')
            .order_by('file_number'),
        )

        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="students_roya.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'رقم الملف', 'تاريخ التسجيل', 'الاسم الأول', 'اسم الأب', 'اسم الجد', 'اسم العائلة',
            'رقم الهوية', 'تاريخ الميلاد', 'الجنس', 'الجنسية', 'الحالة',
            'نوع الإعاقة', 'درجة الإعاقة', 'التشخيص', 'جهة الإحالة',
            'اسم ولي الأمر', 'صلة القرابة', 'رقم هوية ولي الأمر',
            'رقم الجوال', 'جوال إضافي', 'البريد الإلكتروني',
            'العنوان', 'جهة التواصل الرئيسية', 'ملاحظات ولي الأمر',
            'عدد أفراد الأسرة', 'ترتيب المستفيد', 'حالة الوالدين',
            'الدخل الشهري (ريال)', 'الدخل التقريبي', 'نوع السكن',
            'ملاحظات اجتماعية',
        ])

        for student in students:
            guardian = (
                student.guardians.filter(is_primary_contact=True).first()
                or student.guardians.first()
            )
            fam = getattr(student, 'family_info', None)
            writer.writerow([
                student.file_number,
                str(student.registration_date),
                student.first_name,
                student.middle_name,
                student.grandfather_name,
                student.family_name,
                student.national_id,
                str(student.date_of_birth),
                student.get_gender_display(),
                student.nationality,
                student.get_status_display(),
                _disability_display(student),
                student.get_disability_degree_display() if student.disability_degree else '',
                student.diagnosis or '',
                student.get_referral_source_display() if student.referral_source else '',
                guardian.full_name if guardian else '',
                guardian.get_relationship_display() if guardian else '',
                guardian.national_id if guardian else '',
                guardian.phone if guardian else '',
                guardian.phone_alt if guardian else '',
                guardian.email if guardian else '',
                guardian.address if guardian else '',
                'نعم' if (guardian and guardian.is_primary_contact) else 'لا',
                guardian.notes if guardian else '',
                fam.family_size if fam else '',
                fam.sibling_order if fam else '',
                fam.get_parents_status_display() if fam else '',
                fam.monthly_income if fam else '',
                fam.get_income_range_display() if fam else '',
                fam.get_housing_type_display() if fam else '',
                fam.social_notes if fam else '',
            ])

        log_action(request, 'export', None, 'تصدير قائمة الطلاب إلى CSV')
        return response


class StudentImportCsvView(APIView):
    parser_classes = [MultiPartParser, FormParser]
    permission_classes = [CanImport]

    def post(self, request):
        from datetime import date as date_type, datetime

        file = request.FILES.get('file')
        if not file:
            return Response({'detail': 'لم يتم إرفاق ملف.'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            content = file.read().decode('utf-8-sig')
            rows = list(csv.reader(io.StringIO(content)))
        except Exception:
            return Response(
                {'detail': 'تعذّر قراءة الملف. تأكد أنه ملف CSV بترميز UTF-8.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        if len(rows) < 2:
            return Response({'detail': 'الملف فارغ.'}, status=status.HTTP_400_BAD_REQUEST)

        data_rows = rows[1:]
        created, skipped, errors = 0, 0, []

        for i, row in enumerate(data_rows, start=2):
            def _c(idx, r=row):
                return r[idx].strip() if idx < len(r) else ''

            first_name       = _c(0)
            middle_name      = _c(1)
            grandfather_name = _c(2)
            family_name      = _c(3)
            national_id      = _c(4)
            dob_raw          = _c(5)
            gender_raw       = _c(6)
            nationality      = _c(7)

            if not any([first_name, family_name, national_id, dob_raw]):
                continue

            row_errors = []

            if not first_name:
                row_errors.append('الاسم الأول مطلوب')
            if not family_name:
                row_errors.append('اسم العائلة مطلوب')
            if not national_id:
                row_errors.append('رقم الهوية مطلوب')
            elif not national_id.isdigit() or len(national_id) != 10:
                row_errors.append('رقم الهوية يجب أن يكون 10 أرقام')
            if not nationality:
                row_errors.append('الجنسية مطلوبة')

            gender = GENDER_MAP.get(gender_raw)
            if not gender:
                row_errors.append(f'الجنس غير صحيح: "{gender_raw}" — استخدم: ذكر أو أنثى')

            dob = None
            try:
                dob = datetime.strptime(dob_raw, '%Y-%m-%d').date()
                today = date_type.today()
                if dob > today:
                    row_errors.append('تاريخ الميلاد في المستقبل')
                elif dob < today.replace(year=today.year - 100):
                    row_errors.append('تاريخ الميلاد أكبر من 100 سنة')
            except (ValueError, TypeError):
                row_errors.append(f'تاريخ الميلاد غير صحيح: "{dob_raw}" — الصيغة: YYYY-MM-DD')

            full_name_display = f"{first_name} {family_name}".strip()

            if row_errors:
                errors.append({'row': i, 'name': full_name_display or '—', 'errors': row_errors})
                skipped += 1
                continue

            if Student.objects.filter(national_id=national_id).exists():
                errors.append({
                    'row': i,
                    'name': full_name_display,
                    'errors': [f'رقم الهوية {national_id} مسجّل مسبقاً'],
                })
                skipped += 1
                continue

            status_val     = STATUS_MAP.get(_c(8), 'pending') or 'pending'
            reg_date_raw   = _c(9)
            disability_raw = _c(10)
            degree_raw     = _c(11)
            diagnosis      = _c(12)
            edu_raw        = _c(13)
            school_name    = _c(14)
            grade          = _c(15)
            referral_raw   = _c(16)
            notes          = _c(17)

            try:
                reg_date = (
                    datetime.strptime(reg_date_raw, '%Y-%m-%d').date()
                    if reg_date_raw else date_type.today()
                )
            except (ValueError, TypeError):
                reg_date = date_type.today()

            Student.objects.create(
                first_name        = first_name,
                middle_name       = middle_name,
                grandfather_name  = grandfather_name,
                family_name       = family_name,
                national_id       = national_id,
                date_of_birth     = dob,
                gender            = gender,
                nationality       = nationality,
                status            = status_val,
                registration_date = reg_date,
                disability_type   = [DISABILITY_MAP[t.strip()] for t in disability_raw.split(',') if t.strip() in DISABILITY_MAP],
                disability_degree = DEGREE_MAP.get(degree_raw, ''),
                diagnosis         = diagnosis,
                educational_level = EDU_MAP.get(edu_raw, ''),
                school_name       = school_name,
                grade             = grade,
                referral_source   = REFERRAL_MAP.get(referral_raw, ''),
                notes             = notes,
                created_by        = request.user,
            )
            created += 1

        log_action(request, 'create', None, f'استيراد {created} طالب من CSV')
        return Response({'created': created, 'skipped': skipped, 'errors': errors}, status=status.HTTP_200_OK)


class StudentImportCsvTemplateView(APIView):
    permission_classes = [IsAuthenticated]

    def get(self, request):
        response = HttpResponse(content_type='text/csv; charset=utf-8-sig')
        response['Content-Disposition'] = 'attachment; filename="import_template_roya.csv"'

        writer = csv.writer(response)
        writer.writerow([
            'الاسم الأول',
            'اسم الأب',
            'اسم الجد',
            'اسم العائلة',
            'رقم الهوية',
            'تاريخ الميلاد (YYYY-MM-DD)',
            'الجنس (ذكر/أنثى)',
            'الجنسية',
            'الحالة (نشط / في انتظار القبول / غير نشط / خرّيج / موقوف / محوّل)',
            'تاريخ التسجيل (YYYY-MM-DD)',
            'نوع الإعاقة',
            'درجة الإعاقة (بسيطة / متوسطة / شديدة / شديدة جداً)',
            'التشخيص التفصيلي',
            'المستوى التعليمي',
            'اسم المدرسة',
            'الصف / المرحلة',
            'جهة الإحالة',
            'ملاحظات',
        ])
        writer.writerow([
            'محمد', 'أحمد', 'سعد', 'العتيبي', '1234567890', '2010-05-15', 'ذكر', 'سعودي',
            'نشط', '2024-01-10', 'طيف التوحد', 'متوسطة', 'تشخيص طيف التوحد درجة 2',
            'برنامج تربية خاصة', 'مدرسة الأمل', 'الثالث الابتدائي', 'مستشفى / عيادة', '',
        ])
        return response


# ── Attendance ─────────────────────────────────────────────────────────────────
from .models import StudentAttendance
from .serializers import StudentAttendanceSerializer
from django.db.models import Q


class AttendanceListCreateView(generics.ListCreateAPIView):
    serializer_class   = StudentAttendanceSerializer
    permission_classes = [IsAuthenticated]
    pagination_class   = None          # always return full list (no pagination)
    filter_backends    = [DjangoFilterBackend, filters.OrderingFilter]
    filterset_fields   = ['status', 'attendance_date', 'branch', 'guardian_notified']
    ordering_fields    = ['attendance_date', 'created_at']
    ordering           = ['-attendance_date']

    def get_queryset(self):
        student_pk = self.kwargs.get('student_pk')
        if student_pk:
            return (
                StudentAttendance.objects
                .filter(student_id=student_pk)
                .select_related('branch', 'recorded_by', 'updated_by')
            )
        return (
            StudentAttendance.objects
            .select_related('student', 'branch', 'recorded_by', 'updated_by')
        )

    def perform_create(self, serializer):
        from django.db import IntegrityError
        student_pk = self.kwargs.get('student_pk')
        try:
            if student_pk:
                student = get_object_or_404(Student, pk=student_pk)
                # Manual duplicate-date check (UniqueTogetherValidator removed from serializer)
                date = serializer.validated_data.get('attendance_date')
                if StudentAttendance.objects.filter(student=student, attendance_date=date).exists():
                    raise drf_serializers.ValidationError(
                        {'non_field_errors': ['تم تسجيل حضور هذا الطالب مسبقاً في هذا اليوم']}
                    )
                serializer.save(student=student)
            else:
                serializer.save()
        except IntegrityError:
            raise drf_serializers.ValidationError(
                {'non_field_errors': ['تم تسجيل حضور هذا الطالب مسبقاً في هذا اليوم']}
            )


class AttendanceDetailView(generics.RetrieveUpdateDestroyAPIView):
    serializer_class   = StudentAttendanceSerializer
    permission_classes = [IsAuthenticated]

    def get_queryset(self):
        student_pk = self.kwargs.get('student_pk')
        qs = StudentAttendance.objects.select_related('branch', 'recorded_by', 'updated_by')
        if student_pk:
            return qs.filter(student_id=student_pk)
        return qs


class AttendanceSheetView(APIView):
    """
    GET /api/attendance/sheet/?date=YYYY-MM-DD[&branch=ID][&search=text]
    Returns all active students merged with their attendance record for the given date.
    No pagination — designed for the daily attendance register.
    """
    permission_classes = [IsAuthenticated]

    def get(self, request):
        date   = request.query_params.get('date')
        branch = request.query_params.get('branch')
        search = request.query_params.get('search', '').strip()

        if not date:
            return Response({'error': 'date parameter is required'}, status=400)

        students_qs = Student.objects.filter(status='active').select_related('branch')
        if branch:
            students_qs = students_qs.filter(branch_id=branch)
        if search:
            students_qs = students_qs.filter(
                Q(first_name__icontains=search)  |
                Q(family_name__icontains=search) |
                Q(file_number__icontains=search)
            )
        students_qs = students_qs.order_by('first_name', 'family_name')

        # Build attendance map for the date in a single query
        attendance_map = {
            a.student_id: a
            for a in StudentAttendance.objects.filter(
                attendance_date=date,
                student__in=students_qs,
            ).select_related('branch', 'recorded_by')
        }

        result = []
        for s in students_qs:
            att = attendance_map.get(s.id)
            result.append({
                'student_id':   s.id,
                'student_name': s.full_name,
                'file_number':  s.file_number,
                'branch_id':    s.branch_id,
                'branch_name':  s.branch.name if s.branch else None,
                'attendance':   StudentAttendanceSerializer(att).data if att else None,
            })

        return Response(result)
